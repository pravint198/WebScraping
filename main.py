import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
import datetime

wb=Workbook()
ws=wb.active
ws.cell(1,1,'Product')
ws.cell(1,2,'Rating')
ws.cell(1,3,'Price')
row=2
nbPages=int(input("Enter number of pages"))
qString=input("Enter search criteria")
for i in range(1,nbPages+1):
    if i==1:
        page=requests.get(f"https://www.flipkart.com/search?q={qString}&otracker=search&otracker1=search&marketplace=FLIPKART&as-show=on&as=off")
    else:
        page = requests.get(
            f"https://www.flipkart.com/search?q={qString}&otracker=search&otracker1=search&marketplace=FLIPKART&as-show=on&as=off&page={i}")
    soup=BeautifulSoup(page.content,'html.parser')
    #print(soup.prettify())
    products=soup.find_all('div',class_='_1AtVbE col-12-12')


    for product in products:
        product_name=product.find('div',class_='_4rR01T')
        rating=product.find('div',class_='_3LWZlK')
        price=product.find('div',class_='_30jeq3 _1_WHN1')
        if not product_name is None:
            ws.cell(row,1,product_name.text)
            ws.cell(row, 2, rating.text if not rating is None else '')
            ws.cell(row, 3, price.text)
            row=row+1
dt=datetime.datetime.now()
fname="products_"+dt.strftime("%Y%m%d%H%M%S")+".xlsx"
wb.save(fname)